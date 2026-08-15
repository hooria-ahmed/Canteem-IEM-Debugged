"""
CanteenMS — Professional Export Engine
Produces real PDF (ReportLab) and Excel (openpyxl) files.
"""
import io
import csv
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse


# ─── shared helpers ──────────────────────────────────────────────────────────

BRAND     = "Events Up — CanteenMS"
PRIMARY   = (16, 185, 129)   # emerald-500
DARK      = (15,  23,  42)   # slate-900
LIGHT_BG  = (248, 250, 252)  # slate-50
BORDER    = (226, 232, 240)  # slate-200
RED       = (239,  68,  68)


def _xls_styles():
    """Return a dict of ready-made openpyxl style objects."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill

    def _rgb(r, g, b):
        return f"{r:02X}{g:02X}{b:02X}"

    thin = Border(
        left  =Side(style='thin', color=_rgb(*BORDER)),
        right =Side(style='thin', color=_rgb(*BORDER)),
        top   =Side(style='thin', color=_rgb(*BORDER)),
        bottom=Side(style='thin', color=_rgb(*BORDER)),
    )
    return {
        'title'  : Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
        'sub'    : Font(name='Calibri', size=11, color=_rgb(*DARK)),
        'header' : Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        'bold'   : Font(name='Calibri', size=10, bold=True, color=_rgb(*DARK)),
        'normal' : Font(name='Calibri', size=10, color=_rgb(*DARK)),
        'red'    : Font(name='Calibri', size=10, bold=True, color=_rgb(*RED)),
        'green'  : Font(name='Calibri', size=10, bold=True, color=_rgb(*PRIMARY)),
        'fill_hdr': PatternFill('solid', fgColor=_rgb(*DARK)),
        'fill_alt': PatternFill('solid', fgColor=_rgb(*LIGHT_BG)),
        'fill_ttl': PatternFill('solid', fgColor=_rgb(*PRIMARY)),
        'left'   : Alignment(horizontal='left',   vertical='center', wrap_text=True),
        'center' : Alignment(horizontal='center', vertical='center'),
        'right'  : Alignment(horizontal='right',  vertical='center'),
        'border' : thin,
    }


def _xl_header_row(ws, headers, st, row=1, height=28):
    """Write a styled header row."""
    from openpyxl.utils import get_column_letter
    ws.row_dimensions[row].height = height
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = st['header']
        c.fill      = st['fill_hdr']
        c.alignment = st['center']
        c.border    = st['border']
    ws.auto_filter.ref = (
        f"A{row}:{get_column_letter(len(headers))}{row}"
    )


def _xl_data_row(ws, row_idx, values, st, aligns=None):
    """Write one data row with zebra shading."""
    from openpyxl.utils import get_column_letter
    alt = row_idx % 2 == 0
    ws.row_dimensions[row_idx].height = 20
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.font   = st['normal']
        c.border = st['border']
        if alt:
            c.fill = st['fill_alt']
        align_key = aligns[ci - 1] if aligns and ci - 1 < len(aligns) else 'left'
        c.alignment = st[align_key]


def _xl_title_block(ws, st, title, subtitle, num_cols):
    """Merge first row as branded title, second as subtitle."""
    from openpyxl.styles import Alignment as Al
    ws.merge_cells(f"A1:{chr(64+num_cols)}1")
    t = ws['A1']
    t.value     = f"  {BRAND}  —  {title}"
    t.font      = st['title']
    t.fill      = st['fill_ttl']
    t.alignment = Al(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 42

    ws.merge_cells(f"A2:{chr(64+num_cols)}2")
    s = ws['A2']
    s.value     = subtitle
    s.font      = st['sub']
    s.alignment = Al(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 22


def _wb_response(wb, filename):
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


# ─── Finance P&L PDF ──────────────────────────────────────────────────────────

def generate_finance_pdf(context):
    """Professional multi-section P&L PDF with cover page, KPI summary, category breakdown and expense ledger."""
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable, PageBreak, KeepTogether,
    )
    from reportlab.graphics.shapes import Drawing, Rect, String, Line
    from reportlab.graphics import renderPDF
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    W, H = A4
    buf  = io.BytesIO()

    # ── Colour palette ────────────────────────────────────────────────────────
    C_DARK    = colors.Color(0.059, 0.090, 0.165)   # #0F172A slate-900
    C_NAVY    = colors.Color(0.118, 0.227, 0.541)   # #1E3A8A
    C_BLUE    = colors.Color(0.145, 0.392, 0.922)   # #2563EB
    C_GREEN   = colors.Color(0.063, 0.725, 0.506)   # #10B981
    C_GREEN_D = colors.Color(0.024, 0.369, 0.263)   # #065F46
    C_RED     = colors.Color(0.937, 0.267, 0.267)   # #EF4444
    C_GRAY    = colors.Color(0.392, 0.455, 0.545)   # #64748B
    C_LIGHT   = colors.Color(0.973, 0.984, 0.992)   # #F8FAFC
    C_BORDER  = colors.Color(0.886, 0.914, 0.941)   # #E2E8F0
    C_WHITE   = colors.white

    sd = context.get('start_date', '')
    ed = context.get('end_date',   '')
    now_str = datetime.now().strftime('%d %b %Y  %H:%M')

    # ── Paragraph styles ─────────────────────────────────────────────────────
    def PS(name, **kw):
        return ParagraphStyle(name, **kw)

    sH1   = PS('H1',   fontName='Helvetica-Bold', fontSize=26, textColor=C_WHITE,  leading=32)
    sH2   = PS('H2',   fontName='Helvetica-Bold', fontSize=13, textColor=C_DARK,   spaceBefore=20, spaceAfter=8)
    sH3   = PS('H3',   fontName='Helvetica-Bold', fontSize=11, textColor=C_DARK,   spaceAfter=6)
    sSub  = PS('Sub',  fontName='Helvetica',      fontSize=10, textColor=C_GRAY,   spaceAfter=4)
    sBody = PS('Body', fontName='Helvetica',      fontSize=9,  textColor=C_DARK,   leading=14)
    sNote = PS('Note', fontName='Helvetica',      fontSize=8,  textColor=C_GRAY)
    sMono = PS('Mono', fontName='Courier',        fontSize=9,  textColor=C_DARK)
    sCap  = PS('Cap',  fontName='Helvetica-Bold', fontSize=8,  textColor=C_GRAY,
                       textTransform='uppercase', charSpace=0.8)
    sTH   = PS('TH',   fontName='Helvetica-Bold', fontSize=9,  textColor=C_WHITE,  alignment=TA_CENTER)
    sAmt  = PS('Amt',  fontName='Helvetica-Bold', fontSize=9,  textColor=C_DARK,   alignment=TA_RIGHT)
    sAmtG = PS('AmtG', fontName='Helvetica-Bold', fontSize=9,  textColor=C_GREEN,  alignment=TA_RIGHT)
    sAmtR = PS('AmtR', fontName='Helvetica-Bold', fontSize=9,  textColor=C_RED,    alignment=TA_RIGHT)
    sPct  = PS('Pct',  fontName='Helvetica-Bold', fontSize=9,  textColor=C_BLUE,   alignment=TA_RIGHT)

    # ── Data extraction ───────────────────────────────────────────────────────
    total_revenue  = float(context.get('total_revenue',  0))
    total_cogs     = float(context.get('total_cogs',     0))
    gross_profit   = float(context.get('gross_profit',   0))
    total_expenses = float(context.get('total_expenses', 0))
    net_profit     = float(context.get('net_profit',     0))
    net_margin     = float(context.get('net_margin',     0))
    gross_margin   = float(context.get('gross_margin',   0))
    txn_count      = int(context.get('total_txn_count',  0))
    net_positive   = net_profit >= 0

    def rs(v, show_sign=False):
        v = float(v)
        sign = '+ ' if (show_sign and v >= 0) else ('- ' if v < 0 else '')
        return f"{sign}Rs. {abs(v):,.0f}"

    def pct(v):
        return f"{float(v):.1f}%"

    # ── onPage: header + footer drawn on every page ───────────────────────────
    def _on_page(canvas, doc):
        canvas.saveState()
        # Top accent bar
        canvas.setFillColor(C_NAVY)
        canvas.rect(0, H - 1.2*cm, W, 1.2*cm, fill=1, stroke=0)
        canvas.setFillColor(C_GREEN)
        canvas.rect(0, H - 1.2*cm, 0.6*cm, 1.2*cm, fill=1, stroke=0)
        canvas.setFillColor(C_WHITE)
        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(1.2*cm, H - 0.8*cm, 'Events Up — Lucky Cement Canteen')
        canvas.setFont('Helvetica', 9)
        canvas.drawRightString(W - 1.5*cm, H - 0.8*cm, f'Profit & Loss Report  |  {sd} to {ed}')

        # Bottom footer
        canvas.setStrokeColor(C_BORDER)
        canvas.line(1.5*cm, 1.4*cm, W - 1.5*cm, 1.4*cm)
        canvas.setFont('Helvetica', 7.5)
        canvas.setFillColor(C_GRAY)
        canvas.drawString(1.5*cm, 0.85*cm, f'Generated: {now_str}   |   Confidential — For Internal Use Only')
        canvas.drawRightString(W - 1.5*cm, 0.85*cm, f'Page {doc.page}')
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2.0*cm, bottomMargin=2.2*cm,
        title='P&L Report', author='Events Up CanteenMS',
    )

    story = []

    # ════════════════════════════════════════════════════════════════════════════
    # PAGE 1 — COVER
    # ════════════════════════════════════════════════════════════════════════════

    # Full-width cover banner (drawn as a table with background)
    cover_inner = [
        [Paragraph('Events Up', PS('cov1', fontName='Helvetica-Bold', fontSize=9,
                   textColor=C_GREEN, charSpace=2))],
        [Paragraph('Lucky Cement — Korangi Industrial Area, Karachi', PS('cov2',
                   fontName='Helvetica', fontSize=9, textColor=colors.Color(0.7,0.8,1)))],
        [Spacer(1, 0.5*cm)],
        [Paragraph('Profit &amp; Loss<br/>Report', PS('cov3', fontName='Helvetica-Bold',
                   fontSize=34, textColor=C_WHITE, leading=40))],
        [Spacer(1, 0.3*cm)],
        [Paragraph(f'Period: <b>{sd}</b> — <b>{ed}</b>', PS('cov4',
                   fontName='Helvetica', fontSize=12, textColor=colors.Color(0.78, 0.87, 0.98)))],
        [Spacer(1, 0.2*cm)],
        [Paragraph(f'Total Transactions: <b>{txn_count:,}</b>', PS('cov5',
                   fontName='Helvetica', fontSize=10, textColor=colors.Color(0.7, 0.8, 1)))],
    ]
    cover_t = Table(cover_inner, colWidths=[W - 3.6*cm])
    cover_t.setStyle(TableStyle([
        ('BACKGROUND',  (0,0),(-1,-1), C_NAVY),
        ('TOPPADDING',  (0,0),(-1,0),  28),
        ('BOTTOMPADDING',(0,-1),(-1,-1), 32),
        ('LEFTPADDING', (0,0),(-1,-1), 28),
        ('RIGHTPADDING',(0,0),(-1,-1), 20),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[C_NAVY]),
    ]))
    story.append(cover_t)
    story.append(Spacer(1, 0.8*cm))

    # ── 4 KPI tiles ──────────────────────────────────────────────────────────
    def kpi_tile(label, value, sub, color=C_DARK):
        return Table(
            [[Paragraph(label.upper(), PS('kl', fontName='Helvetica-Bold', fontSize=7,
                        textColor=C_GRAY, charSpace=0.5, spaceAfter=4))],
             [Paragraph(value, PS('kv', fontName='Helvetica-Bold', fontSize=18,
                        textColor=color, leading=22))],
             [Paragraph(sub, PS('ks', fontName='Helvetica', fontSize=8, textColor=C_GRAY))]],
            colWidths=[(W - 3.6*cm) / 4 - 0.3*cm]
        )

    tiles_data = [[
        kpi_tile('Gross Revenue',     f"Rs. {total_revenue:,.0f}",
                 f"{txn_count} orders", C_DARK),
        kpi_tile('Cost of Goods',     f"Rs. {total_cogs:,.0f}",
                 f"{gross_margin:.1f}% gross margin", C_GRAY),
        kpi_tile('Gross Profit',      f"Rs. {gross_profit:,.0f}",
                 'After COGS deduction', C_NAVY),
        kpi_tile('Net Profit',        f"Rs. {abs(net_profit):,.0f}",
                 f"{net_margin:.1f}% net margin {'▲' if net_positive else '▼'}",
                 C_GREEN if net_positive else C_RED),
    ]]
    col_w = (W - 3.6*cm) / 4
    tiles = Table(tiles_data, colWidths=[col_w]*4, hAlign='LEFT')
    tiles.setStyle(TableStyle([
        ('BOX',         (0,0),(-1,-1), 1, C_BORDER),
        ('LINEAFTER',   (0,0),(2,0),   0.5, C_BORDER),
        ('BACKGROUND',  (0,0),(-1,-1), C_WHITE),
        ('TOPPADDING',  (0,0),(-1,-1), 16),
        ('BOTTOMPADDING',(0,0),(-1,-1),14),
        ('LEFTPADDING', (0,0),(-1,-1), 16),
        ('RIGHTPADDING',(0,0),(-1,-1), 8),
        ('VALIGN',      (0,0),(-1,-1), 'TOP'),
    ]))
    story.append(tiles)
    story.append(Spacer(1, 0.7*cm))

    # ════════════════════════════════════════════════════════════════════════════
    # P&L STATEMENT TABLE
    # ════════════════════════════════════════════════════════════════════════════
    story.append(Paragraph('Profit &amp; Loss Statement', sH2))

    def pl_row(label, amount, style=sBody, bg=C_WHITE, bold=False, indent=0):
        indent_str = '&nbsp;' * indent
        lbl_style = PS('lr', fontName='Helvetica-Bold' if bold else 'Helvetica',
                       fontSize=9, textColor=C_DARK, leading=14)
        amt_style = PS('ar', fontName='Helvetica-Bold' if bold else 'Helvetica',
                       fontSize=9, textColor=C_DARK, alignment=TA_RIGHT, leading=14)
        return [Paragraph(indent_str + label, lbl_style),
                Paragraph(amount, amt_style)]

    pl_data = [
        # Header
        [Paragraph('Description', sTH), Paragraph('Amount (Rs.)', sTH)],
        # Revenue
        pl_row('REVENUE', '', bold=True),
        pl_row('Gross Sales Revenue', f"Rs. {total_revenue:,.0f}", indent=4),
        pl_row('Total Revenue', f"Rs. {total_revenue:,.0f}", bold=True),
        # COGS
        pl_row('COST OF GOODS SOLD', '', bold=True),
        pl_row('Inventory / Ingredient Cost', f"(Rs. {total_cogs:,.0f})", indent=4),
        pl_row('Gross Profit', f"Rs. {gross_profit:,.0f}", bold=True),
        pl_row('Gross Margin %', pct(gross_margin), bold=False),
        # Expenses
        pl_row('OPERATING EXPENSES', '', bold=True),
        pl_row('Total Operating Expenses', f"(Rs. {total_expenses:,.0f})", indent=4),
        # Net
        pl_row('NET PROFIT / (LOSS)', rs(net_profit), bold=True),
        pl_row('Net Profit Margin', pct(net_margin), bold=False),
    ]

    pl_table = Table(pl_data, colWidths=[11*cm, 5.2*cm])
    pl_style = [
        # Header row
        ('BACKGROUND', (0,0),(-1,0), C_DARK),
        ('FONTNAME',   (0,0),(-1,0), 'Helvetica-Bold'),
        ('TEXTCOLOR',  (0,0),(-1,0), C_WHITE),
        ('FONTSIZE',   (0,0),(-1,-1), 9),
        ('PADDING',    (0,0),(-1,-1), 9),
        ('ALIGN',      (1,0),(1,-1),  'RIGHT'),
        ('GRID',       (0,0),(-1,-1), 0.4, C_BORDER),
        ('ROWBACKGROUNDS', (0,1),(-1,-1), [C_WHITE, C_LIGHT]),
        # Section header rows (Revenue, COGS, Expenses)
        ('BACKGROUND', (0,1),(-1,1), colors.Color(0.93, 0.96, 1.0)),
        ('BACKGROUND', (0,4),(-1,4), colors.Color(0.93, 0.96, 1.0)),
        ('BACKGROUND', (0,8),(-1,8), colors.Color(0.93, 0.96, 1.0)),
        # Gross profit row
        ('BACKGROUND', (0,6),(-1,6), colors.Color(0.92, 1.0, 0.96)),
        ('FONTNAME',   (0,6),(-1,6), 'Helvetica-Bold'),
        # Net profit row
        ('BACKGROUND', (0,10),(-1,10),
         colors.Color(0.90, 1.0, 0.94) if net_positive else colors.Color(1.0, 0.92, 0.92)),
        ('FONTNAME',   (0,10),(-1,10), 'Helvetica-Bold'),
        ('TEXTCOLOR',  (1,10),(1,10), C_GREEN if net_positive else C_RED),
    ]
    pl_table.setStyle(TableStyle(pl_style))
    story.append(pl_table)
    story.append(Spacer(1, 0.8*cm))

    # ════════════════════════════════════════════════════════════════════════════
    # REVENUE BY CATEGORY
    # ════════════════════════════════════════════════════════════════════════════
    sales_by_cat = context.get('sales_by_category', [])
    if sales_by_cat:
        story.append(KeepTogether([
            Paragraph('Revenue by Category', sH2),
        ]))

        # Table
        cat_hdr = [Paragraph(h, sTH) for h in ['Category', 'Revenue (Rs.)', 'Share %', 'Visual']]
        cat_rows = [cat_hdr]
        for cat in sales_by_cat:
            pct_val = float(cat['percent'])
            amt_val = float(cat['amount'])
            # Mini bar (drawn as a table cell background trick via a filled cell)
            bar_w = max(pct_val / 100, 0.02)
            bar_cell = Table(
                [['']], colWidths=[6*cm * bar_w + 0.01],
                rowHeights=[10],
            )
            bar_cell.setStyle(TableStyle([
                ('BACKGROUND', (0,0),(0,0), C_GREEN),
                ('PADDING', (0,0),(-1,-1), 0),
            ]))
            cat_rows.append([
                Paragraph(cat['name'], sBody),
                Paragraph(f"Rs. {amt_val:,.0f}", sAmtG),
                Paragraph(f"{pct_val:.1f}%", sPct),
                bar_cell,
            ])

        cat_table = Table(cat_rows, colWidths=[6.5*cm, 3.5*cm, 2*cm, 4.2*cm])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,0), C_DARK),
            ('FONTSIZE',      (0,0),(-1,-1), 9),
            ('PADDING',       (0,0),(-1,-1), 9),
            ('ALIGN',         (1,0),(2,-1),  'RIGHT'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [C_WHITE, C_LIGHT]),
            ('GRID',          (0,0),(-1,-1), 0.4, C_BORDER),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 0.8*cm))

    # ════════════════════════════════════════════════════════════════════════════
    # EXPENSE LEDGER — new page
    # ════════════════════════════════════════════════════════════════════════════
    recent_exp = list(context.get('recent_expenses', []))
    if recent_exp:
        story.append(PageBreak())
        story.append(Paragraph('Expense Ledger', sH2))
        story.append(Paragraph(
            f'Showing {len(recent_exp)} most recent expenses for the selected period.',
            sSub))
        story.append(Spacer(1, 0.3*cm))

        exp_hdr = [Paragraph(h, sTH) for h in ['Date', 'Title', 'Category', 'Amount (Rs.)', 'Recorded By']]
        exp_rows = [exp_hdr]
        exp_total = 0.0
        for e in recent_exp:
            amt = float(e.amount)
            exp_total += amt
            exp_rows.append([
                Paragraph(str(e.expense_date), sBody),
                Paragraph(e.title or '—', sBody),
                Paragraph(e.category.name if e.category else 'Uncategorized', sBody),
                Paragraph(f"Rs. {amt:,.0f}", PS('ea', fontName='Helvetica',
                          fontSize=9, alignment=TA_RIGHT, textColor=C_DARK)),
                Paragraph(e.recorded_by.full_name if e.recorded_by else 'Admin', sBody),
            ])
        # Total row
        exp_rows.append([
            Paragraph('', sBody), Paragraph('', sBody),
            Paragraph('Total', PS('et', fontName='Helvetica-Bold', fontSize=9, textColor=C_DARK)),
            Paragraph(f"Rs. {exp_total:,.0f}", PS('etv', fontName='Helvetica-Bold',
                      fontSize=9, alignment=TA_RIGHT, textColor=C_RED)),
            Paragraph('', sBody),
        ])

        exp_table = Table(exp_rows, colWidths=[2.8*cm, 4.5*cm, 3*cm, 3.2*cm, 2.7*cm])
        exp_table.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,0),  C_DARK),
            ('FONTSIZE',      (0,0),(-1,-1), 9),
            ('PADDING',       (0,0),(-1,-1), 9),
            ('ROWBACKGROUNDS',(0,1),(-1,-2), [C_WHITE, C_LIGHT]),
            ('GRID',          (0,0),(-1,-1), 0.4, C_BORDER),
            ('ALIGN',         (3,0),(3,-1),  'RIGHT'),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            # Total row
            ('BACKGROUND',    (0,-1),(-1,-1), colors.Color(1.0, 0.94, 0.94)),
            ('FONTNAME',      (0,-1),(-1,-1), 'Helvetica-Bold'),
            ('LINEABOVE',     (0,-1),(-1,-1), 1, C_RED),
        ]))
        story.append(exp_table)

    # ── Build ─────────────────────────────────────────────────────────────────
    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)

    fname = f"PL_Report_{sd}_to_{ed}.pdf"
    resp  = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp




# ─── Finance Excel (full workbook) ───────────────────────────────────────────

def generate_finance_excel(context):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    wb = openpyxl.Workbook()
    st = _xls_styles()

    start = context.get('start_date', '')
    end   = context.get('end_date',   '')
    total_revenue  = float(context.get('total_revenue',  0))
    total_cogs     = float(context.get('total_cogs',     0))
    gross_profit   = float(context.get('gross_profit',   0))
    total_expenses = float(context.get('total_expenses', 0))
    net_profit     = float(context.get('net_profit',     0))
    net_margin     = float(context.get('net_margin',     0))
    gross_margin   = float(context.get('gross_margin',   0))
    txn_count      = int(context.get('total_txn_count',  0))

    # ── Sheet 1: P&L Summary ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "P&L Summary"
    ws1.sheet_view.showGridLines = False
    _xl_title_block(ws1, st, "Profit & Loss Report", f"Period: {start} to {end}", 3)

    metrics = [
        ('Gross Revenue',       total_revenue,   '#CURRENCY#', st['green']),
        ('Cost of Goods Sold',  -total_cogs,     '#CURRENCY#', st['normal']),
        ('Gross Profit',        gross_profit,    '#CURRENCY#', st['bold']),
        ('Gross Margin %',      gross_margin/100,'#PCT#',      st['normal']),
        ('Operating Expenses',  -total_expenses, '#CURRENCY#', st['normal']),
        ('Net Profit / Loss',   net_profit,      '#CURRENCY#', st['green'] if net_profit >= 0 else st['red']),
        ('Net Margin %',        net_margin/100,  '#PCT#',      st['normal']),
        ('Orders Processed',    txn_count,       '#INT#',      st['normal']),
    ]

    ws1.cell(row=3, column=1, value='Metric').font = st['bold']
    ws1.cell(row=3, column=2, value='Value').font  = st['bold']
    ws1.cell(row=3, column=3, value='Notes').font  = st['bold']
    for c in [1,2,3]:
        ws1.cell(row=3, column=c).fill      = st['fill_hdr']
        ws1.cell(row=3, column=c).font      = st['header']
        ws1.cell(row=3, column=c).alignment = st['center']
        ws1.cell(row=3, column=c).border    = st['border']
    ws1.row_dimensions[3].height = 24

    for i, (label, val, fmt, font) in enumerate(metrics, 4):
        alt = i % 2 == 0
        cl = ws1.cell(row=i, column=1, value=label)
        cv = ws1.cell(row=i, column=2, value=val)
        cn = ws1.cell(row=i, column=3, value='')
        for c in [cl, cv, cn]:
            c.border    = st['border']
            c.alignment = st['left']
            if alt: c.fill = st['fill_alt']
        cl.font = font
        cv.font = font
        cv.alignment = st['right']
        if fmt == '#CURRENCY#':
            cv.number_format = '"Rs. "#,##0.00;[Red]"- Rs. "#,##0.00'
        elif fmt == '#PCT#':
            cv.number_format = '0.00%'
        else:
            cv.number_format = '#,##0'
        ws1.row_dimensions[i].height = 20

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 20
    ws1.freeze_panes = 'A4'

    # ── Sheet 2: Revenue by Category ─────────────────────────────────────────
    ws2 = wb.create_sheet("By Category")
    ws2.sheet_view.showGridLines = False
    _xl_title_block(ws2, st, "Revenue by Category", f"Period: {start} to {end}", 3)
    headers2 = ['Category', 'Revenue (Rs.)', 'Share %']
    _xl_header_row(ws2, headers2, st, row=3)
    for i, cat in enumerate(context.get('sales_by_category', []), 4):
        _xl_data_row(ws2, i,
            [cat['name'], float(cat['amount']), float(cat['percent'])/100],
            st, aligns=['left','right','right'])
        ws2.cell(row=i, column=2).number_format = '"Rs. "#,##0'
        ws2.cell(row=i, column=3).number_format = '0.0%'
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 16
    ws2.freeze_panes = 'A4'

    # ── Sheet 3: Transactions ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Transactions")
    ws3.sheet_view.showGridLines = False
    _xl_title_block(ws3, st, "Sales Transactions", f"Period: {start} to {end}", 6)
    h3 = ['Txn No', 'Date', 'Time', 'Payment', 'Amount (Rs.)', 'Status']
    _xl_header_row(ws3, h3, st, row=3)
    txns = context.get('_txns_qs', [])
    for i, t in enumerate(txns, 4):
        _xl_data_row(ws3, i, [
            t.transaction_no,
            str(t.sale_date),
            t.sale_time.strftime('%H:%M:%S') if t.sale_time else '',
            t.payment_method.upper(),
            float(t.total_amount),
            t.status.upper(),
        ], st, aligns=['left','center','center','center','right','center'])
        ws3.cell(row=i, column=5).number_format = '"Rs. "#,##0.00'
    for ci, w in enumerate([20,14,12,14,18,14], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = 'A4'

    # ── Sheet 4: Expenses ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Expenses")
    ws4.sheet_view.showGridLines = False
    _xl_title_block(ws4, st, "Expense Ledger", f"Period: {start} to {end}", 5)
    h4 = ['Date', 'Title', 'Category', 'Amount (Rs.)', 'Recorded By']
    _xl_header_row(ws4, h4, st, row=3)
    for i, e in enumerate(context.get('_expenses_qs', []), 4):
        _xl_data_row(ws4, i, [
            str(e.expense_date),
            e.title or '—',
            e.category.name if e.category else '—',
            float(e.amount),
            e.recorded_by.full_name if e.recorded_by else 'Admin',
        ], st, aligns=['center','left','left','right','left'])
        ws4.cell(row=i, column=4).number_format = '"Rs. "#,##0.00'
    for ci, w in enumerate([14,28,20,18,20], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.freeze_panes = 'A4'

    fname = f"FinanceReport_{start}_to_{end}.xlsx"
    return _wb_response(wb, fname)


# ─── Audit Log CSV ────────────────────────────────────────────────────────────

def generate_audit_csv(logs_qs):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = (
        f'attachment; filename="AuditLog_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
    )
    writer = csv.writer(resp)
    writer.writerow(['Timestamp', 'Operator', 'Action Code', 'Entity Type',
                     'Entity ID', 'IP Address', 'Old Values', 'New Values'])
    for log in logs_qs:
        writer.writerow([
            log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            log.user.full_name if log.user else '—',
            log.action_code,
            log.entity_type,
            log.entity_id or '',
            log.ip_address or '',
            str(log.old_values or ''),
            str(log.new_values or ''),
        ])
    return resp


# ─── Order History Excel ──────────────────────────────────────────────────────

def generate_orders_excel(txns_qs, date_filter=''):
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Order History"
    ws.sheet_view.showGridLines = False
    st  = _xls_styles()

    subtitle = f"Filtered by: {date_filter}" if date_filter else "All Records"
    _xl_title_block(ws, st, "Order History", subtitle, 7)

    headers = ['Txn No', 'Date', 'Time', 'Cashier', 'Payment', 'Amount (Rs.)', 'Status']
    _xl_header_row(ws, headers, st, row=3)

    for i, t in enumerate(txns_qs, 4):
        _xl_data_row(ws, i, [
            t.transaction_no,
            str(t.sale_date),
            t.sale_time.strftime('%H:%M') if t.sale_time else '',
            t.cashier.full_name if t.cashier else '—',
            t.payment_method.upper(),
            float(t.total_amount),
            t.status.upper(),
        ], st, aligns=['left','center','center','left','center','right','center'])
        ws.cell(row=i, column=6).number_format = '"Rs. "#,##0'

    for ci, w in enumerate([20, 14, 10, 20, 12, 18, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    fname = f"Orders_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _wb_response(wb, fname)
