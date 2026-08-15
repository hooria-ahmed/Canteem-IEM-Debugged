import os
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from core.models import Dish, DishCategory, DishIngredient, RawMaterial, RawMaterialCategory


def clean_decimal(value, default=None):
    try:
        if value is None:
            return default
        return Decimal(str(value).replace(',', '').strip())
    except (InvalidOperation, ValueError, TypeError):
        return default


def map_units(value):
    if not value:
        return 'piece'
    unit = str(value).strip().lower().replace('(', '').replace(')', '')
    aliases = {
        'kg': 'kg',
        'g': 'g',
        'gm': 'g',
        'gram': 'g',
        'grams': 'g',
        'litre': 'litre',
        'liter': 'litre',
        'ltr': 'litre',
        'l': 'litre',
        'ml': 'ml',
        'pcs': 'piece',
        'pc': 'piece',
        'piece': 'piece',
        'pieces': 'piece',
        'nos': 'piece',
        'slice': 'piece',
        'ord': 'piece',
        'small': 'piece',
        'pkt': 'packet',
        'pck': 'packet',
        'pack': 'packet',
        'packet': 'packet',
        'bag': 'bag',
        'doz': 'dozen',
        'dozen': 'dozen',
        'bunch': 'bunch',
        'gadi': 'bunch',
        'bottle': 'bottle',
        'btl': 'bottle',
        'bot': 'bottle',
        'blk': 'block',
        'block': 'block',
        'tin': 'tin',
        'crate': 'crate',
        'glass': 'glass',
    }
    return aliases.get(unit, 'piece')


class Command(BaseCommand):
    help = 'Import recipe ingredients and ingredient costs from the recipes workbook.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='data/Recipes MIS - SITE & BBQ.xlsx',
            help='Path to the recipes Excel file',
        )

    def handle(self, *args, **kwargs):
        file_path = kwargs['file']
        if not os.path.isabs(file_path):
            file_path = os.path.join(settings.BASE_DIR, file_path)
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True)
            sheet = workbook['Recipes']
        except Exception as exc:
            self.stdout.write(self.style.ERROR(f'Error loading workbook: {exc}'))
            return

        # Parse first, then write in batches. The original importer issued several
        # SELECT/INSERT/UPDATE queries for every one of the 4k+ spreadsheet rows.
        rows = []
        skipped = 0
        categories_needed = {}
        dish_meta = {}
        material_meta = {}
        recipe_links = {}

        for row in sheet.iter_rows(min_row=3, values_only=True):
            dish_name = row[0]
            category_name = row[1]
            ingredient_name = row[3]
            qty = clean_decimal(row[4])
            mapped_unit = map_units(row[5])
            site_rate = clean_decimal(row[10])

            if not dish_name or not ingredient_name or qty is None or qty <= 0:
                skipped += 1
                continue

            dish_name = str(dish_name).strip()
            ingredient_name = str(ingredient_name).strip()
            category_name = str(category_name).strip() if category_name else 'Uncategorized'
            dish_key = dish_name.casefold()
            material_key = ingredient_name.casefold()
            category_key = category_name.casefold()

            categories_needed[category_key] = category_name
            dish_meta.setdefault(dish_key, {'name': dish_name, 'category_key': category_key})

            material = material_meta.setdefault(
                material_key,
                {'name': ingredient_name, 'unit': mapped_unit, 'site_rate': Decimal('0.00')},
            )
            # Use the latest explicit workbook rate, matching the old sequential
            # update behavior, but never erase a real cost because a cell is blank.
            if site_rate is not None and site_rate >= 0:
                material['site_rate'] = site_rate

            # Duplicate dish/material rows exist in the source. The final row for
            # a pair wins, matching update_or_create behavior without thousands of
            # database round trips.
            recipe_links[(dish_key, material_key)] = {
                'qty': qty,
                'unit': mapped_unit,
            }
            rows.append((dish_key, material_key))

        default_rm_category, _ = RawMaterialCategory.objects.get_or_create(name='General')
        now = timezone.now()

        with transaction.atomic():
            existing_categories = {c.name.casefold(): c for c in DishCategory.objects.all()}
            missing_categories = [
                DishCategory(name=name, description='Imported from recipe workbook')
                for key, name in categories_needed.items()
                if key not in existing_categories
            ]
            if missing_categories:
                DishCategory.objects.bulk_create(missing_categories, batch_size=250, ignore_conflicts=True)
            category_map = {c.name.casefold(): c for c in DishCategory.objects.all()}

            # Preserve existing selling prices. The recipe workbook contains
            # ingredient rates, not menu selling prices.
            dish_map = {}
            for dish in Dish.objects.order_by('id'):
                dish_map.setdefault(dish.name.strip().casefold(), dish)

            new_dishes = []
            for key, meta in dish_meta.items():
                if key not in dish_map:
                    new_dishes.append(Dish(
                        name=meta['name'],
                        category=category_map[meta['category_key']],
                        selling_price=Decimal('0.01'),
                        cost_price=Decimal('0.00'),
                        is_available=False,
                    ))
            if new_dishes:
                Dish.objects.bulk_create(new_dishes, batch_size=250)

            dish_map = {}
            for dish in Dish.objects.order_by('id'):
                dish_map.setdefault(dish.name.strip().casefold(), dish)

            dishes_to_update = []
            for key, meta in dish_meta.items():
                dish = dish_map[key]
                if dish.category_id is None:
                    dish.category = category_map[meta['category_key']]
                    dish.updated_at = now
                    dishes_to_update.append(dish)
            if dishes_to_update:
                Dish.objects.bulk_update(dishes_to_update, ['category', 'updated_at'], batch_size=250)

            material_map = {}
            for material in RawMaterial.objects.order_by('id'):
                material_map.setdefault(material.name.strip().casefold(), material)

            new_materials = []
            for key, meta in material_meta.items():
                if key not in material_map:
                    new_materials.append(RawMaterial(
                        name=meta['name'],
                        category=default_rm_category,
                        units=meta['unit'],
                        cost_per_unit=meta['site_rate'],
                    ))
            if new_materials:
                RawMaterial.objects.bulk_create(new_materials, batch_size=250)

            material_map = {}
            for material in RawMaterial.objects.order_by('id'):
                material_map.setdefault(material.name.strip().casefold(), material)

            materials_to_update = []
            for key, meta in material_meta.items():
                material = material_map[key]
                if material.cost_per_unit != meta['site_rate']:
                    material.cost_per_unit = meta['site_rate']
                    material.updated_at = now
                    materials_to_update.append(material)
            if materials_to_update:
                RawMaterial.objects.bulk_update(
                    materials_to_update,
                    ['cost_per_unit', 'updated_at'],
                    batch_size=250,
                )

            relevant_dish_ids = [dish_map[key].id for key in dish_meta]
            existing_links = {
                (link.dish_id, link.raw_material_id): link
                for link in DishIngredient.objects.filter(dish_id__in=relevant_dish_ids)
            }
            links_to_create = []
            links_to_update = []

            for (dish_key, material_key), data in recipe_links.items():
                dish = dish_map[dish_key]
                material = material_map[material_key]
                pair = (dish.id, material.id)
                link = existing_links.get(pair)
                if link is None:
                    links_to_create.append(DishIngredient(
                        dish=dish,
                        raw_material=material,
                        quantity_required=data['qty'],
                        unit=data['unit'],
                    ))
                elif link.quantity_required != data['qty'] or link.unit != data['unit']:
                    link.quantity_required = data['qty']
                    link.unit = data['unit']
                    link.updated_at = now
                    links_to_update.append(link)

            if links_to_create:
                DishIngredient.objects.bulk_create(links_to_create, batch_size=500)
            if links_to_update:
                DishIngredient.objects.bulk_update(
                    links_to_update,
                    ['quantity_required', 'unit', 'updated_at'],
                    batch_size=500,
                )

        self.stdout.write(self.style.SUCCESS(f'Imported/updated {len(rows)} recipe lines.'))
        self.stdout.write(
            f'Created {len(new_dishes)} unavailable dishes awaiting real menu prices; '
            f'{len(recipe_links)} unique recipe links are represented.'
        )
        if skipped:
            self.stdout.write(self.style.WARNING(f'Skipped {skipped} empty or invalid rows.'))
