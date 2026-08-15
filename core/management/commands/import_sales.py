from datetime import datetime
from decimal import Decimal, InvalidOperation
import os

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from core.models import Dish, MealSession, SaleItem, SalesTransaction, User


def clean_number(value):
    if value is None:
        return None
    try:
        if isinstance(value, str):
            value = value.replace(',', '').strip()
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def parse_sale_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def parse_slot(value):
    text = str(value or '').strip().lower()
    if 'midnight' in text:
        return 'snack', '00:00:00'
    if 'breakfast' in text:
        return 'breakfast', '08:00:00'
    if 'lunch' in text:
        return 'lunch', '13:00:00'
    if 'afternoon' in text:
        return 'snack', '17:00:00'
    if 'dinner' in text:
        return 'dinner', '20:00:00'
    return None, '12:00:00'


class Command(BaseCommand):
    help = 'Import historical sales from the Daily Sale Excel workbook.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', type=str, default='data/daily_sales.xlsx',
            help='Path to the sales Excel file',
        )

    def handle(self, *args, **kwargs):
        file_path = kwargs['file']
        if not os.path.isabs(file_path):
            file_path = os.path.join(settings.BASE_DIR, file_path)
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        if MealSession.objects.count() == 0:
            for name, start, end in [
                ('breakfast', '07:00', '11:00'),
                ('lunch', '12:00', '16:00'),
                ('snack', '16:00', '19:00'),
                ('dinner', '19:00', '23:59'),
            ]:
                MealSession.objects.create(name=name, start_time=start, end_time=end, is_active=True)

        cashier, _ = User.objects.get_or_create(
            user_name='import_user',
            defaults={
                'password_hash': '', 'full_name': 'Data Import User',
                'role': 'cashier', 'is_active': False,
            },
        )

        sessions = {s.name: s for s in MealSession.objects.filter(is_active=True)}
        default_session = next(iter(sessions.values()), None)
        if default_session is None:
            self.stdout.write(self.style.ERROR('No active meal session is available.'))
            return

        wb = load_workbook(file_path, data_only=True, read_only=True)
        sheet = next((wb[name] for name in wb.sheetnames if 'sale' in name.lower()), wb.active)
        self.stdout.write(f'Using sheet: {sheet.title}')

        records = []
        skipped = 0
        block_starts = (0, 8)  # A:G and I:O
        skip_words = ('TOTAL', 'GRAND', 'CASH FROM', 'CREDIT', 'DEPARTMENT', 'ITEM NAME')

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=15, values_only=True), start=2):
            for block_no, base in enumerate(block_starts, start=1):
                item_name = row[base + 1] if len(row) > base + 1 else None
                if not item_name:
                    continue
                item_name = str(item_name).strip()
                if not item_name or any(word in item_name.upper() for word in skip_words):
                    skipped += 1
                    continue

                sale_date = parse_sale_date(row[base + 2])
                qty = clean_number(row[base + 4])
                rate = clean_number(row[base + 5])
                amount = clean_number(row[base + 6])
                if sale_date is None or qty is None or rate is None or amount is None:
                    skipped += 1
                    continue
                if qty <= 0 or rate < 0 or amount < 0:
                    skipped += 1
                    continue

                session_name, sale_time = parse_slot(row[base + 3])
                meal_session = sessions.get(session_name) or default_session
                txn_id = f"IMP-{sale_date.strftime('%Y%m%d')}-{row_index:04d}-{block_no}"
                records.append({
                    'txn_id': txn_id,
                    'item_name': item_name,
                    'dish_key': item_name.casefold(),
                    'sale_date': sale_date,
                    'sale_time': sale_time,
                    'qty': qty,
                    'rate': rate,
                    'amount': amount,
                    'meal_session': meal_session,
                    'order_key': (sale_date, row_index, block_no),
                })

        existing_ids = set(
            SalesTransaction.objects.filter(transaction_no__startswith='IMP-')
            .values_list('transaction_no', flat=True)
        )
        pending = [r for r in records if r['txn_id'] not in existing_ids]
        duplicates = len(records) - len(pending)

        if not pending:
            self.stdout.write(self.style.SUCCESS('Imported 0 sales records.'))
            if skipped:
                self.stdout.write(self.style.WARNING(f'Skipped {skipped} header/summary/invalid rows.'))
            if duplicates:
                self.stdout.write(self.style.WARNING(f'Skipped {duplicates} records already imported.'))
            return

        # Use the latest dated sales row as the current menu rate for each dish.
        latest_by_dish = {}
        for record in records:
            current = latest_by_dish.get(record['dish_key'])
            if current is None or record['order_key'] > current['order_key']:
                latest_by_dish[record['dish_key']] = record

        existing_dishes = {d.name.strip().casefold(): d for d in Dish.objects.all()}
        new_dishes = []
        for key, record in latest_by_dish.items():
            if key not in existing_dishes:
                new_dishes.append(Dish(
                    name=record['item_name'],
                    selling_price=record['rate'],
                    cost_price=Decimal('0.00'),
                    is_available=True,
                ))

        created_dishes = 0
        updated_prices = 0
        with transaction.atomic():
            if new_dishes:
                Dish.objects.bulk_create(new_dishes, batch_size=250)
                created_dishes = len(new_dishes)

            dish_map = {d.name.strip().casefold(): d for d in Dish.objects.all()}
            to_update = []
            for key, record in latest_by_dish.items():
                dish = dish_map[key]
                # Avoid violating the stored margin constraint if a manually set
                # cost_price is higher than an old historical sales rate.
                if record['rate'] >= dish.cost_price and (
                    dish.selling_price != record['rate'] or not dish.is_available
                ):
                    dish.selling_price = record['rate']
                    dish.is_available = True
                    to_update.append(dish)
            if to_update:
                Dish.objects.bulk_update(to_update, ['selling_price', 'is_available', 'updated_at'], batch_size=250)
                updated_prices = len(to_update)

            txns = [SalesTransaction(
                transaction_no=r['txn_id'],
                cashier=cashier,
                meal_session=r['meal_session'],
                sale_date=r['sale_date'],
                sale_time=r['sale_time'],
                sub_total=r['amount'],
                discount_amount=Decimal('0.00'),
                total_amount=r['amount'],
                amount_received=r['amount'],
                change_returned=Decimal('0.00'),
                payment_method='cash',
                status='completed',
                notes='Imported historical sale; stock was not deducted by this import.',
            ) for r in pending]
            SalesTransaction.objects.bulk_create(txns, batch_size=250)
            txn_map = {t.transaction_no: t for t in txns}

            items = []
            for r in pending:
                qty_int = max(1, int(r['qty'].to_integral_value()))
                items.append(SaleItem(
                    transaction=txn_map[r['txn_id']],
                    dish=dish_map[r['dish_key']],
                    dish_name=dish_map[r['dish_key']].name,
                    quantity=qty_int,
                    unit_price=r['rate'],
                    line_total=r['amount'],
                ))
            SaleItem.objects.bulk_create(items, batch_size=250)

        self.stdout.write(self.style.SUCCESS(f'Imported {len(pending)} sales records.'))
        self.stdout.write(f'Created {created_dishes} dishes; updated {updated_prices} menu prices.')
        if skipped:
            self.stdout.write(self.style.WARNING(f'Skipped {skipped} header/summary/invalid rows.'))
        if duplicates:
            self.stdout.write(self.style.WARNING(f'Skipped {duplicates} records already imported.'))
